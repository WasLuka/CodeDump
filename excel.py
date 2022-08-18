from openpyxl import Workbook
wb = Workbook()
abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
# grab the active worksheet
ws = wb.active
i = 0
for x in abc:
    for y in range(1, 27):
        ws[f'{x}{y}'] = y + 26*i
    i += 1

ws['A27'] = '=AVERAGE(A1:A26)'

wb.save("sample.xlsx")
